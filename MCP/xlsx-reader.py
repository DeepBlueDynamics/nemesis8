#!/usr/bin/env python3
"""
MCP: xlsx-reader

Utilities for reading XLSX files (questionnaires/assessments).
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

from mcp.server.fastmcp import FastMCP

try:
    import openpyxl
    XLSX_AVAILABLE = True
except Exception:
    openpyxl = None
    XLSX_AVAILABLE = False

mcp = FastMCP("xlsx-reader")


def _load_wb(xlsx_path: Path):
    if not XLSX_AVAILABLE:
        raise RuntimeError("openpyxl not installed in MCP venv.")
    return openpyxl.load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)


@mcp.tool()
async def xlsx_list_sheets(xlsx_path: str) -> Dict[str, object]:
    """List sheet names in an XLSX file."""
    path = Path(xlsx_path)
    if not path.exists():
        return {"success": False, "error": "file_not_found", "path": str(path)}
    try:
        wb = _load_wb(path)
        return {"success": True, "path": str(path), "sheets": wb.sheetnames}
    except Exception as e:
        return {"success": False, "error": str(e), "path": str(path)}


@mcp.tool()
async def xlsx_read_range(
    xlsx_path: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
) -> Dict[str, object]:
    """Read a cell range (e.g., A1:C10) from a sheet."""
    path = Path(xlsx_path)
    if not path.exists():
        return {"success": False, "error": "file_not_found", "path": str(path)}
    try:
        wb = _load_wb(path)
        if sheet_name not in wb.sheetnames:
            return {"success": False, "error": "sheet_not_found", "sheet": sheet_name}
        ws = wb[sheet_name]
        data = []
        for row in ws[start_cell:end_cell]:
            data.append([cell.value for cell in row])
        return {
            "success": True,
            "path": str(path),
            "sheet": sheet_name,
            "start_cell": start_cell,
            "end_cell": end_cell,
            "rows": data,
        }
    except Exception as e:
        return {"success": False, "error": str(e), "path": str(path)}


@mcp.tool()
async def xlsx_read_table(
    xlsx_path: str,
    sheet_name: Optional[str] = None,
) -> Dict[str, object]:
    """Read the used range of a sheet."""
    path = Path(xlsx_path)
    if not path.exists():
        return {"success": False, "error": "file_not_found", "path": str(path)}
    try:
        wb = _load_wb(path)
        target_sheet = sheet_name or wb.sheetnames[0]
        if target_sheet not in wb.sheetnames:
            return {"success": False, "error": "sheet_not_found", "sheet": target_sheet}
        ws = wb[target_sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return {
            "success": True,
            "path": str(path),
            "sheet": target_sheet,
            "rows": rows,
            "row_count": len(rows),
        }
    except Exception as e:
        return {"success": False, "error": str(e), "path": str(path)}


@mcp.tool()
async def xlsx_find_questions(
    xlsx_path: str,
    sheet_name: Optional[str] = None,
    max_rows: int = 500,
) -> Dict[str, object]:
    """Heuristic scan for question-like rows (Q1, Question, etc.)."""
    path = Path(xlsx_path)
    if not path.exists():
        return {"success": False, "error": "file_not_found", "path": str(path)}
    try:
        wb = _load_wb(path)
        target_sheet = sheet_name or wb.sheetnames[0]
        if target_sheet not in wb.sheetnames:
            return {"success": False, "error": "sheet_not_found", "sheet": target_sheet}
        ws = wb[target_sheet]
        hits = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i > max_rows:
                break
            row_text = " ".join([str(v) for v in row if v is not None]).lower()
            if any(tok in row_text for tok in ["question", "q1", "q2", "q3", "q4", "q5"]):
                hits.append({"row": i, "values": list(row)})
        return {
            "success": True,
            "path": str(path),
            "sheet": target_sheet,
            "matches": hits,
            "match_count": len(hits),
        }
    except Exception as e:
        return {"success": False, "error": str(e), "path": str(path)}


if __name__ == "__main__":
    mcp.run()
